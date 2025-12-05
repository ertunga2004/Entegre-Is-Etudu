import re
from typing import Dict, List, Tuple
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# Series ve ChartLines buraya eklendi
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel




# ------------------------- Yardımcı Fonksiyonlar -------------------------

def _read_csv_safe(path: str) -> pd.DataFrame:
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame()

def _ensure_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df

def _is_genel_tekrar(param_code) -> bool:
    if pd.isna(param_code):
        return False
    x = str(param_code).strip().lower().replace(" ", "")
    return x.startswith("geneltekrar")

def _extract_indis(param_code) -> str:
    """ParametreKodu başı: harf(ler) + bitişik rakam(lar) (örn. A, A1, B)."""
    if pd.isna(param_code):
        return ""
    s = str(param_code).strip()
    m = re.match(r'^([A-Za-zÇĞİÖŞÜçğıöşü]+[0-9]*)', s)
    return m.group(1) if m else ""
def _parse_predecessors(value) -> List[int]:
    """
    '1', '1,2', '1 2', '1;2' gibi yazılmış öncül listelerini
    [1, 2] şeklinde int listesine çevirir.
    """
    if pd.isna(value):
        return []
    s = str(value).strip()
    if not s:
        return []
    # ; / boşluk gibi ayraçları virgüna çevir
    s = s.replace(";", ",").replace("/", ",")
    tokens = re.split(r"[,\s]+", s)
    result = []
    for t in tokens:
        t = t.strip()
        if not t:
            continue
        try:
            result.append(int(t))
        except ValueError:
            # sayı değilse yok say
            pass
    return result

def _split_model_and_sekans(model_tipi: str) -> Tuple[str, str]:
    """'AAA-BBB' -> ('AAA','BBB'); '-' yoksa sekans boş kalır."""
    if pd.isna(model_tipi):
        return "", ""
    s = str(model_tipi)
    if "-" in s:
        left, right = s.split("-", 1)
        return left.strip(), right.strip()
    return s.strip(), ""

def _pick_label(row: pd.Series, choices):
    """choices = [(kolon_adı, görünen_etiket), ...] -> ilk TRUE olanı döndür."""
    for col, lbl in choices:
        try:
            v = row.get(col, 0)
        except Exception:
            v = 0
        if pd.notna(v) and str(v).strip() not in ["0", "0.0", "", "False", "NaN"]:
            return lbl
    return "-"

def _multi_labels(row: pd.Series, choices):
    labels = []
    for col, lbl in choices:
        v = row.get(col, 0)
        if pd.notna(v) and str(v).strip() not in ["0", "0.0", "", "False", "NaN"]:
            labels.append(lbl)
    return labels if labels else ["-"]


# ------------------------- Raporlayıcı Sınıf -------------------------

class MostExcelJobReport:
    """
    Her 'İş Adı' için ayrı çalışma sayfası açar ve iki blok üretir:

    SOL (A..G): Westinghouse blokları (AnalizID başına 13 satır):
        A: Adım Adı (13 satır birleştirilmiş, ortalı)
        B: Performans Faktörleri
        C: Performans Faktörleri Değerleri
        D: İşin Zorluk Faktörleri
        E: Zorluk Faktörleri Değerleri
        F: Genel Koşullar (başlığın altından itibaren, her değer ayrı satır; yoksa "-")
        G: Koruyucu Ekipmanlar (başlığın altından itibaren, her değer ayrı satır; yoksa "-")

    H–I: Boş sütunlar.

    SAĞ (J..P): MOST özeti:
        "Adım Adı | Model | Sekans | Genel Tekrar İndis | İndis | İndis Tekrarı | Açıklama"
        - Model: 'ModelTipi' '-' öncesi
        - Sekans: 'ModelTipi' '-' sonrası
        - Genel Tekrar: 'Genel Tekrar' satırındaki TekrarSayisi, yoksa 1
        - İndis: ParametreKodu başı (örn. A, A1)
        - İndis Tekrarı: TekrarSayisi
        - Açıklama: SecilenDeger
        - Hücre birleştirme: Aynı AnalizID bloğunda Adım Adı + Model + Sekans + Genel Tekrar İndis
    """

    # MOST başlıkları
    MOST_HEADERS = ["Adım Adı", "Model", "Sekans", "Genel Tekrar İndis",
                    "İndis", "İndis Tekrarı", "Açıklama"]

    # Westinghouse başlıkları (blok üst satırı)
    WH_HEADERS = ["Performans Faktörleri", "Performans Faktörleri Değerleri",
                  "İşin Zorluk Faktörleri", "Zorluk Faktörleri Değerleri",
                  "Genel Koşullar", "Koruyucu Ekipmanlar"]

    # Sabit satır sıraları
    WH_PF_NAMES = ["Yetenek", "Çaba", "Çalışma Koşulları", "Tutarlılık"]
    WH_ZF_NAMES = ["Kişisel Gereksinimler", "Fiziksel Çaba", "Düşünsel Çaba",
                   "Çalışma Pozisyonu", "Atmosfer", "Isı", "Gürültü"]

    def __init__(self, paths: Dict[str, str], selected_job_ids=None):
        """
        paths:
            {
              "is_adimlari": ".../Is_Adimlari.csv",
              "basic_analiz": ".../Basic_Most_Analizleri.csv",
              "basic_detay":  ".../Basic_Most_Detaylari.csv",
              "maxi_analiz":  ".../Maxi_Most_Analizleri.csv",
              "maxi_detay":   ".../Maxi_Most_Detaylari.csv",
              "mini_analiz":  ".../Mini_Most_Analizleri.csv",
              "mini_detay":   ".../Mini_Most_Detaylari.csv",
              "westinghouse": ".../Westinghouse_Analizleri.csv",
            }
        """
        self.paths = paths
        self.selected_job_ids = [int(j) for j in selected_job_ids] if selected_job_ids else []


        # Stil
        self.header_fill = PatternFill("solid", fgColor="D9E1F2")
        self.header_font = Font(bold=True)
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )


        # Veri setleri
        self.is_adimlari = pd.DataFrame()
        self.analiz_all = pd.DataFrame(columns=["AnalizID", "AdımID", "ModelTipi"])
        self.detay_all  = pd.DataFrame(columns=["AnalizID", "ParametreKodu", "SecilenDeger", "TekrarSayisi"])
        self.west       = pd.DataFrame()
        self.job_to_sheet: Dict[str, str] = {}


    # ------------ Yardımcı: WH kolonlarını normalize et ------------
    def _normalize_west_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df

        # Beklediğimiz ana kolonlar
        required = ["AnalizID", "JobID", "AdımID", "NormalZaman", "StandartZaman"]
        for col in required:
            if col not in df.columns:
                df[col] = pd.NA

        # AnalizID / AdımID sayısal olsun
        for col in ["AnalizID", "AdımID"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Zaman kolonlarını float'a çevir
        for col in ["NormalZaman", "StandartZaman"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        return df

        # ------------ Veri Yükleme ------------

    def _load(self):
        dfs = {k: _read_csv_safe(v) for k, v in self.paths.items()}

        # İş tablosu
        # DataManager'ın oluşturduğu kolonlarla hizalıyoruz:
        # ['JobID', 'AdımID', 'İş Adı', 'Adım Adı', 'Öncül Adım', 'DegerTuru', 'İş İstasyonu']
        self.is_adimlari = _ensure_cols(
            dfs.get("is_adimlari", pd.DataFrame()),
            ["JobID", "İş Adı", "AdımID", "Adım Adı", "Öncül Adım", "DegerTuru", "İş İstasyonu"]
        )

        # Zaman etüdü (opsiyonel) – Timestamp'i de alalım, Ana Sayfa'da kullanacağız
        self.zaman_etudu = _ensure_cols(
            dfs.get("zaman_etudu", pd.DataFrame()),
            ["JobID", "AdımID", "Sure", "Timestamp"]
        )

        # (Bundan sonrası – analiz_all, detay_all, west – olduğu gibi kalsın)

        # Analizler (Basic/Maxi/Mini birleştirme) – ToplamSaniye'yi de alıyoruz
        analiz_frames = []
        for k in ["basic_analiz", "maxi_analiz", "mini_analiz"]:
            df = dfs.get(k, pd.DataFrame()).copy()
            if df.empty:
                continue
            df = _ensure_cols(df, ["AnalizID", "AdımID", "ModelTipi", "ToplamSaniye"])
            if df["ModelTipi"].isna().all():
                fallback = {"basic_analiz": "BasicMOST",
                            "maxi_analiz": "MaxiMOST",
                            "mini_analiz": "MiniMOST"}[k]
                df["ModelTipi"] = fallback
            analiz_frames.append(df[["AnalizID", "AdımID", "ModelTipi", "ToplamSaniye"]])

        self.analiz_all = (pd.concat(analiz_frames, ignore_index=True)
                           if analiz_frames else
                           pd.DataFrame(columns=["AnalizID", "AdımID", "ModelTipi", "ToplamSaniye"]))

        # Detaylar (Basic/Maxi/Mini birleştirme)
        detay_frames = []
        for k in ["basic_detay", "maxi_detay", "mini_detay"]:
            df = dfs.get(k, pd.DataFrame()).copy()
            if df.empty:
                continue
            df = _ensure_cols(df, ["AnalizID", "ParametreKodu", "SecilenDeger", "TekrarSayisi"])
            detay_frames.append(df[["AnalizID", "ParametreKodu", "SecilenDeger", "TekrarSayisi"]])

        self.detay_all = (pd.concat(detay_frames, ignore_index=True)
                          if detay_frames else
                          pd.DataFrame(columns=["AnalizID", "ParametreKodu", "SecilenDeger", "TekrarSayisi"]))

        # Westinghouse
        self.west = dfs.get("westinghouse", pd.DataFrame()).copy()
        if not self.west.empty:
            self.west = self._normalize_west_columns(self.west)
            # AnalizID/AdımID sayısal — güvenli eşleşme
            for col in ["AnalizID", "AdımID"]:
                if col in self.west.columns:
                    self.west[col] = pd.to_numeric(self.west[col], errors="coerce")

    # ------------ Excel Yardımcıları ------------

    def _autosize(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                except Exception:
                    val = ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    def _draw_frame(self, ws, sr, er, sc, ec):
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                ws.cell(row=r, column=c).border = self.thin_border


    def _write_sop_band(self, ws, start_row, start_col, method_label, adim_sayisi):

        sc = start_col
        ec = start_col + 8
        sr = start_row

        grey = PatternFill("solid", fgColor="D9D9D9")

        # VA / NVAN / NVA için özel renkler
        va_fill   = PatternFill("solid", fgColor="00B050")  # yeşil
        nvan_fill = PatternFill("solid", fgColor="FFFF00")  # sarı
        nva_fill  = PatternFill("solid", fgColor="FF0000")  # kırmızı

        # 1) Başlık
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr+2, end_column=ec)
        cell = ws.cell(row=sr, column=sc, value="STANDART OPERASYON PROSEDÜRÜ (SOP)")
        cell.alignment = self.center
        self._draw_frame(ws, sr, sr+2, sc, ec)

        # 2) PROSES
        ws.merge_cells(start_row=sr+3, start_column=sc, end_row=sr+3, end_column=ec)
        cell = ws.cell(row=sr+3, column=sc, value=f"PROSES: {ws.title}")
        cell.alignment = self.center
        self._draw_frame(ws, sr+3, sr+3, sc, ec)

        # 3) Uygulanan Yöntem
        ws.merge_cells(start_row=sr+4, start_column=sc, end_row=sr+4, end_column=ec)
        cell = ws.cell(row=sr+4, column=sc, value=f"UYGULANAN YÖNTEM : {method_label}")
        cell.alignment = self.center
        self._draw_frame(ws, sr+4, sr+4, sc, ec)

        # 4) VAR satırı
        var = sr+6
        ws.merge_cells(start_row=var, start_column=sc,   end_row=var, end_column=sc+4)
        ws.merge_cells(start_row=var, start_column=sc+5, end_row=var, end_column=sc+6)
        self._draw_frame(ws, var, var, sc, ec)

        # 5) CT / TT satırı
        ct = sr+8
        # CT label
        ws.merge_cells(start_row=ct, start_column=sc, end_row=ct, end_column=sc+2)
        cell = ws.cell(row=ct, column=sc, value="ÇEVRİM SÜRESİ (CT)")
        cell.fill = grey
        cell.alignment = self.center

        # CT value + sn
        ws.cell(row=ct, column=sc+3)          # değer hücresi
        cell = ws.cell(row=ct, column=sc+4, value="sn")
        cell.alignment = self.center

        # TT label (2 hücre)
        ws.merge_cells(start_row=ct, start_column=sc+5, end_row=ct, end_column=sc+6)
        cell = ws.cell(row=ct, column=sc+5, value="TAKT ZAMANI (TT)")
        cell.fill = grey
        cell.alignment = self.center

        # TT value + sn
        ws.cell(row=ct, column=sc+7)          # değer hücresi
        cell = ws.cell(row=ct, column=sc+8, value="sn")
        cell.alignment = self.center

        self._draw_frame(ws, ct, ct, sc, ec)

        # 6) Operasyon başlıkları
        h = sr + 11
        n = max(10, adim_sayisi)

        # No
        cell = ws.cell(row=h, column=sc, value="No")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = grey

        # ID
        cell = ws.cell(row=h, column=sc+1, value="ID")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = grey

        # OPERASYON ADIMLARI  (C ve D sütunları birleşik)
        ws.merge_cells(start_row=h, start_column=sc+2, end_row=h, end_column=sc+3)
        cell = ws.cell(row=h, column=sc+2, value="OPERASYON ADIMLARI")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = grey

        # ÖNCÜL
        cell = ws.cell(row=h, column=sc+4, value="ÖNCÜL")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = grey

        # VA / NVAN / NVA başlıkları (renkli)
        cell = ws.cell(row=h, column=sc+5, value="VA (sn)")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = va_fill

        cell = ws.cell(row=h, column=sc+6, value="NVAN (sn)")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = nvan_fill

        cell = ws.cell(row=h, column=sc+7, value="NVA (sn)")
        cell.font = self.header_font
        cell.alignment = self.center
        cell.fill = nva_fill

        # Başlık satırının çerçevesi
        self._draw_frame(ws, h, h, sc, ec)

        # 7) Operasyon satırları (C ve D her satırda birleşik)
               # 7) Boş satırlar
        for r in range(h + 1, h + 1 + n):
            # OPERASYON ADIMLARI sütunlarını (C ve D) her satırda birleştir
            ws.merge_cells(start_row=r, start_column=sc+2, end_row=r, end_column=sc+3)
            self._draw_frame(ws, r, r, sc, ec)

        # 8) TOPLAM / SWIP / GÜVENLİK / DOKÜMAN satır indexlerini,
        #   çizilen operasyon satırı sayısına göre ayarla
        t = h + 1 + n           # TOPLAM SÜRE satırı

                # 8a) CT hücresine formül: TOPLAM VA + TOPLAM NVAN + TOPLAM NVA
        ct_row = ct
        toplam_row = t  # "toplam süre" satırı

        va_top   = ws.cell(row=toplam_row, column=sc + 5).coordinate
        nvan_top = ws.cell(row=toplam_row, column=sc + 6).coordinate
        nva_top  = ws.cell(row=toplam_row, column=sc + 7).coordinate

        ct_cell = ws.cell(row=ct_row, column=sc + 3)
        ct_cell.value = f"={va_top}+{nvan_top}+{nva_top}"
        ct_cell.alignment = self.center

        s = t + 2               # SWIP bloğu başı
        g = s + 4               # GÜVENLİK bloğu başı
        d = g + 3               # DOKÜMAN satırı

        # 8) TOPLAM SÜRE – tek satır, senin ekran görüntündeki gibi
        # Solda: No + ID + OPERASYON ADIMLARI + ÖNCÜL (sc..sc+4) birleşik hücre
        ws.merge_cells(start_row=t, start_column=sc,
                       end_row=t, end_column=sc + 4)
        cell = ws.cell(row=t, column=sc, value="toplam süre")
        cell.alignment = self.center

        # Sağda: VA / NVAN / NVA sütunlarında AYRI AYRI toplam hücreleri
        for col_offset in (5, 6, 7):  # VA, NVAN, NVA sütunları
            total_cell = ws.cell(row=t, column=sc + col_offset)
            total_cell.alignment = self.center  # buraya toplam değer/ formül gelecek

        # Bu satırın çerçevesi
        self._draw_frame(ws, t, t, sc, ec)

        # 9) SWIP BLOĞU
        ws.merge_cells(start_row=s, start_column=sc, end_row=s+3, end_column=sc+2)
        cell = ws.cell(row=s, column=sc, value="SWIP")
        cell.font = self.header_font
        cell.alignment = self.center
        self._draw_frame(ws, s, s+3, sc, sc+2)

        ws.merge_cells(start_row=s, start_column=sc+3, end_row=s+3, end_column=ec)
        cell = ws.cell(
            row=s,
            column=sc+3,
            value="SWIP : İşlem görmüş maksimum stoğun adedi.\n"
                  "Eğer işlenen adet belirlenen değerin üzerindeyse çalışmayı durdur."
        )
        cell.alignment = self.left
        self._draw_frame(ws, s, s+3, sc+3, ec)

        # 10) GÜVENLİK BLOĞU
        ws.merge_cells(start_row=g, start_column=sc, end_row=g+2, end_column=ec)
        cell = ws.cell(row=g, column=sc, value="GÜVENLİK")
        cell.font = self.header_font
        cell.alignment = self.center
        self._draw_frame(ws, g, g+2, sc, ec)

        # 11) DOKÜMAN SATIRI
        ws.merge_cells(start_row=d, start_column=sc, end_row=d, end_column=ec)
        cell = ws.cell(
            row=d,
            column=sc,
            value="DOKÜMAN NO: SOP.İŞ  |  REV NO: 0  |  REV TARİHİ: --.--.----  |  İLK YAYIN TARİHİ: --.--.----"
        )
        cell.alignment = self.center
        self._draw_frame(ws, d, d, sc, ec)

        return d


        # 10) GÜVENLİK BLOĞU
        ws.merge_cells(start_row=g, start_column=sc, end_row=g+2, end_column=ec)
        cell = ws.cell(row=g, column=sc, value="GÜVENLİK")
        cell.font = self.header_font
        cell.alignment = self.center
        self._draw_frame(ws, g, g+2, sc, ec)

        # 11) DOKÜMAN SATIRI
        ws.merge_cells(start_row=d, start_column=sc, end_row=d, end_column=ec)
        cell = ws.cell(
            row=d,
            column=sc,
            value="DOKÜMAN NO: SOP.İŞ  |  REV NO: 0  |  REV TARİHİ: --.--.----  |  İLK YAYIN TARİHİ: --.--.----"
        )
        cell.alignment = self.center
        self._draw_frame(ws, d, d, sc, ec)

        return d
    
    def _add_percent_chart_for_sop_band(self, ws, start_row, start_col, adim_sayisi, anchor_offset_cols=9):
        """
        Tek bir SOP bandı (Zaman Etüdü / Westinghouse / MOST) için
        VA–NVAN–NVA yüzdelerini gösteren %100 yığılmış sütun grafiği ekler.

        start_row : SOP bandının başladığı satır (bizde 2)
        start_col : SOP bandının başladığı sütun (Zaman=1, West=11, MOST=21)
        """
        sr = start_row
        sc = start_col

        # _write_sop_band içindeki ile aynı mantık:
        # h = sr + 11 -> VA/NVAN/NVA başlık satırı
        # n = max(10, adim_sayisi)
        # t = h + 1 + n -> "toplam süre" satırı
        h = sr + 11
        n = max(10, adim_sayisi)
        t = h + 1 + n  # toplam satır

        # Toplam hücreleri (VA / NVAN / NVA)
        va_cell   = ws.cell(row=t, column=sc + 5)
        nvan_cell = ws.cell(row=t, column=sc + 6)
        nva_cell  = ws.cell(row=t, column=sc + 7)

        # Tamamen boşsa grafik üretme
        if all(c.value in (None, "", 0) for c in (va_cell, nvan_cell, nva_cell)):
            return

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "percentStacked"
        chart.overlap = 100

        # Yüzdelik eksen görünümü
        chart.y_axis.number_format = "0%"   # 0% 10% 20% ... 100%
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.y_axis.majorUnit = 0.1        # 10'ar 10'ar artsın
        chart.y_axis.tickLblPos = "nextTo"  # etiketleri göster

        chart.y_axis.number_format = "0%"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1

        # Veriler: sadece toplamların olduğu satır
        data = Reference(
            ws,
            min_col=sc + 5,   # VA
            max_col=sc + 7,   # NVA
            min_row=t,
            max_row=t
        )
        chart.add_data(data, titles_from_data=False)
        from openpyxl.chart.label import DataLabelList  # en üst importlarda da var olsun

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True          # sadece sayıyı yaz
        chart.dataLabels.showSerName = False     # VA (sn) yazma
        chart.dataLabels.showCatName = False     # "toplam süre" yazma
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showPercent = False
        

        # --- Seri isimleri (VA / NVAN / NVA) ---
        for i, s in enumerate(chart.series):
            label = SeriesLabel()
            label.v = ws.cell(row=h, column=sc + 5 + i).value
            s.title = label

            # Renkler: VA=yeşil, NVAN=sarı, NVA=kırmızı
            if i == 0:      # VA
                s.graphicalProperties.solidFill = "00B050"   # yeşil
            elif i == 1:    # NVAN
                s.graphicalProperties.solidFill = "FFFF00"   # sarı
            else:           # NVA
                s.graphicalProperties.solidFill = "FF0000"   # kırmızı


        # --- Veri etiketleri: 12.0 / 8.0 / 22.0 barların ortasında gözüksün ---
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True      # saniye değerlerini yaz
        # Eğer istersen yüzdelik de yazsın:
        # chart.dataLabels.showPercent = True
        chart.dataLabels.position = "ctr"    # etiketler dilimin ORTASINDA

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        

        # Seri isimleri: başlık satırındaki "VA (sn) / NVAN (sn) / NVA (sn)"
        for i, s in enumerate(chart.series):


            for i, s in enumerate(chart.series):
                series_label = SeriesLabel()
                series_label.v = ws.cell(row=h, column=sc + 5 + i).value
                s.title = series_label

        # Kategori etiketi: "toplam süre" yazan hücre (tek kategori)
        cats = Reference(
            ws,
            min_col=sc,       # "toplam süre" yazan sütun
            max_col=sc,
            min_row=t,
            max_row=t
        )
        chart.set_categories(cats)

        # Boyut – ince uzun görünmesi için
        chart.height = 10
        chart.width = 4

        # Grafiği SOP bandının sağına yerleştir
        anchor_col = sc + anchor_offset_cols   # istiyorsan 8, 9, 10 ile oynayabilirsin
        anchor_cell = f"{get_column_letter(anchor_col)}{h}"
        ws.add_chart(chart, anchor_cell)

    def _add_sop_percent_charts(self, ws, start_row, adim_sayisi):
        """
        Üstteki 3 SOP bandının (Zaman Etüdü, Westinghouse, MOST)
        yanına yüzde dağılım grafiklerini ekler.
        """
        # SOL: Zaman Etüdü blok (start_col = 1)
        self._add_percent_chart_for_sop_band(ws, start_row=start_row, start_col=1,  adim_sayisi=adim_sayisi)

        # ORTA: Westinghouse blok (start_col = 11)
        self._add_percent_chart_for_sop_band(ws, start_row=start_row, start_col=11, adim_sayisi=adim_sayisi)

        # SAĞ: MOST blok (start_col = 21)
        self._add_percent_chart_for_sop_band(ws, start_row=start_row, start_col=21, adim_sayisi=adim_sayisi)











    def _fill_zaman_sop_from_time_study(
        self,
        ws,
        start_row: int,
        adim_sayisi: int,
        job_df: pd.DataFrame,
        zaman_map: Dict
    ) -> None:
        """
        Sadece SOL taraftaki (Zaman Etüdü) SOP tablosunu doldurur.
        Kaynaklar:
          - Is_Adimlari.csv  -> AdımID, Adım Adı, Öncül Adım, DegerTuru
          - Zaman_Etudu.csv  -> AdımID bazında Sure (ortalama süre)
        """

        # SOP bandının konumu (sol blok)
        sc = 1                    # başlangıç sütunu
        sr = start_row            # SOP'un başladığı satır (bizde 2)
        ec = sc + 8

        # _write_sop_band içindeki aynı hesaplar:
        h = sr + 11               # tablo başlık satırı (No, ID, ...)
        n = max(10, adim_sayisi)  # en az 10 satır
        data_start = h + 1        # ilk operasyon satırı
        data_end = h + n          # son operasyon satırı
        toplam_row = h + 1 + n    # "TOPLAM SÜRE" satırı

        # İş adımlarını hazırlayalım
        steps = job_df.dropna(subset=["AdımID"]).copy()
        if steps.empty:
            return

        # AdımID int olsun ve sıralı gidelim
        steps["AdımID"] = steps["AdımID"].astype(int)
        steps = steps.sort_values("AdımID")

        # DegerTuru -> VA / NVAN / NVA sınıflaması
        def _classify(deger):
            if not isinstance(deger, str):
                return ""
            d = deger.strip().upper()
            if d in ("KD", "VA"):
                return "VA"
            if d in ("ZGK", "NVAN", "NVAA", "NV-AN"):
                return "NVAN"
            if d in ("KDZ", "NVA"):
                return "NVA"
            return ""

        current_row = data_start
        no = 1

        for _, r in steps.iterrows():
            if current_row > data_end:
                break  # güvenlik

            adim_id = r.get("AdımID")
            op_name = r.get("Adım Adı")
            oncul = r.get("Öncül Adım")
            cat = _classify(r.get("DegerTuru"))

            # Zaman Etüdü'nden süre (yoksa 0)
            sure = float(zaman_map.get(adim_id, 0.0) or 0.0)

            va = nv_an = nva = 0.0
            if cat == "VA":
                va = sure
            elif cat == "NVAN":
                nv_an = sure
            elif cat == "NVA":
                nva = sure

            # Hücrelere yaz
            ws.cell(row=current_row, column=sc,     value=no).alignment = self.center    # No
            ws.cell(row=current_row, column=sc + 1, value=adim_id).alignment = self.center
            ws.cell(row=current_row, column=sc + 2, value=op_name).alignment = self.left
            ws.cell(row=current_row, column=sc + 4, value=oncul).alignment = self.center
            ws.cell(row=current_row, column=sc + 5, value=round(va, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 6, value=round(nv_an, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 7, value=round(nva, 4)).alignment = self.center

            # Kenarlık
            for col in range(sc, sc + 8):
                ws.cell(row=current_row, column=col).border = self.thin_border

            no += 1
            current_row += 1

        # --- TOPLAM SÜRE satırı için formüller ---
        # (VA, NVAN, NVA sütunlarının toplamı)
        if data_start <= data_end:
            # VA toplamı
           # VA toplamı
            first_va = ws.cell(row=data_start, column=sc + 5).coordinate
            last_va  = ws.cell(row=data_end,   column=sc + 5).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 5,
                value=f"=SUM({first_va}:{last_va})"
            ).alignment = self.center

            # NVAN toplamı
            first_nv = ws.cell(row=data_start, column=sc + 6).coordinate
            last_nv  = ws.cell(row=data_end,   column=sc + 6).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 6,
                value=f"=SUM({first_nv}:{last_nv})"
            ).alignment = self.center

            # NVA toplamı
            first_nva = ws.cell(row=data_start, column=sc + 7).coordinate
            last_nva  = ws.cell(row=data_end,   column=sc + 7).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 7,
                value=f"=SUM({first_nva}:{last_nva})"
            ).alignment = self.center
            # --- ÇEVRİM SÜRESİ (CT) hücresini doldur ---
            ct_row = sr + 8          # CT satırı
            ct_col = sc + 3          # CT değer hücresi (sn yazan hücrenin solu)

            va_total_addr   = ws.cell(row=toplam_row, column=sc + 5).coordinate
            nvan_total_addr = ws.cell(row=toplam_row, column=sc + 6).coordinate
            nva_total_addr  = ws.cell(row=toplam_row, column=sc + 7).coordinate

            ct_cell = ws.cell(
                row=ct_row,
                column=ct_col,
                value=f"=SUM({va_total_addr},{nvan_total_addr},{nva_total_addr})"
            )
            ct_cell.alignment = self.center

            # --- TAKT ZAMANI (TT) hücresini hazırla ---
            tt_row = ct_row
            tt_col = sc + 7          # TT değer hücresi (sn yazan hücrenin solu)

            tt_cell = ws.cell(row=tt_row, column=tt_col)
            # Burayı istersen sabit yapabilirsin, örn:
            # tt_cell.value = 60    # 60 sn gibi
            # Şimdilik boş bırakıyorum, kullanıcı Excel'den girsin:
            tt_cell.value = None
            tt_cell.alignment = self.center

    def _fill_west_sop_from_westhouse(
        self,
        ws,
        start_row: int,
        adim_sayisi: int,
        job_df: pd.DataFrame,
        west_map: Dict
    ) -> None:
        """
        Ortadaki (Westinghouse) SOP tablosunu doldurur.
        Kaynaklar:
          - Is_Adimlari.csv  -> AdımID, Adım Adı, Öncül Adım, DegerTuru
          - Westinghouse_Analizleri.csv -> AdımID bazında Standart Süre (west_map)
        """

        # Westinghouse SOP bandı orta blok: start_col = 11
        sc = 11                   # başlangıç sütunu (Westinghouse SOP)
        sr = start_row            # SOP'un başladığı satır (bizde 2)
        ec = sc + 8

        # _write_sop_band içindeki aynı hesaplar:
        h = sr + 11               # tablo başlık satırı (No, ID, ...)
        n = max(10, adim_sayisi)  # en az 10 satır
        data_start = h + 1        # ilk operasyon satırı
        data_end = h + n          # son operasyon satırı
        toplam_row = h + 1 + n    # "TOPLAM SÜRE" satırı

        # İş adımlarını hazırlayalım
        steps = job_df.dropna(subset=["AdımID"]).copy()
        if steps.empty:
            return

        # AdımID int olsun ve sıralı gidelim
        steps["AdımID"] = steps["AdımID"].astype(int)
        steps = steps.sort_values("AdımID")

        # DegerTuru -> VA / NVAN / NVA sınıflaması (Zaman Etüdü ile bire bir aynı)
        def _classify(deger):
            if not isinstance(deger, str):
                return ""
            d = deger.strip().upper()
            if d in ("KD", "VA"):
                return "VA"
            if d in ("ZGK", "NVAN", "NVAA", "NV-AN"):
                return "NVAN"
            if d in ("KDZ", "NVA"):
                return "NVA"
            return ""

        current_row = data_start
        no = 1

        for _, r in steps.iterrows():
            if current_row > data_end:
                break  # güvenlik

            adim_id = r.get("AdımID")
            op_name = r.get("Adım Adı")
            oncul = r.get("Öncül Adım")
            cat = _classify(r.get("DegerTuru"))

            # Westinghouse analizinden standart süre (yoksa 0)
            sure = float(west_map.get(adim_id, 0.0) or 0.0)

            va = nv_an = nva = 0.0
            if cat == "VA":
                va = sure
            elif cat == "NVAN":
                nv_an = sure
            elif cat == "NVA":
                nva = sure

            # Hücrelere yaz (ortadaki SOP bandı)
            ws.cell(row=current_row, column=sc,     value=no).alignment = self.center    # No
            ws.cell(row=current_row, column=sc + 1, value=adim_id).alignment = self.center
            ws.cell(row=current_row, column=sc + 2, value=op_name).alignment = self.left
            ws.cell(row=current_row, column=sc + 4, value=oncul).alignment = self.center
            ws.cell(row=current_row, column=sc + 5, value=round(va, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 6, value=round(nv_an, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 7, value=round(nva, 4)).alignment = self.center

            # Kenarlık
            for col in range(sc, sc + 8):
                ws.cell(row=current_row, column=col).border = self.thin_border

            no += 1
            current_row += 1

        # --- TOPLAM SÜRE satırı için formüller (VA / NVAN / NVA) ---
        if data_start <= data_end:
            # VA toplamı
            first_va = ws.cell(row=data_start, column=sc + 5).coordinate
            last_va  = ws.cell(row=data_end,   column=sc + 5).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 5,
                value=f"=SUM({first_va}:{last_va})"
            ).alignment = self.center

            # NVAN toplamı
            first_nv = ws.cell(row=data_start, column=sc + 6).coordinate
            last_nv  = ws.cell(row=data_end,   column=sc + 6).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 6,
                value=f"=SUM({first_nv}:{last_nv})"
            ).alignment = self.center

            # NVA toplamı
            first_nva = ws.cell(row=data_start, column=sc + 7).coordinate
            last_nva  = ws.cell(row=data_end,   column=sc + 7).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 7,
                value=f"=SUM({first_nva}:{last_nva})"
            ).alignment = self.center
            # --- ÇEVRİM SÜRESİ (CT) hücresini doldur ---
            ct_row = sr + 8
            ct_col = sc + 3

            va_total_addr   = ws.cell(row=toplam_row, column=sc + 5).coordinate
            nvan_total_addr = ws.cell(row=toplam_row, column=sc + 6).coordinate
            nva_total_addr  = ws.cell(row=toplam_row, column=sc + 7).coordinate

            ct_cell = ws.cell(
                row=ct_row,
                column=ct_col,
                value=f"=SUM({va_total_addr},{nvan_total_addr},{nva_total_addr})"
            )
            ct_cell.alignment = self.center

            # --- TAKT ZAMANI (TT) hücresini hazırla ---
            tt_row = ct_row
            tt_col = sc + 7

            tt_cell = ws.cell(row=tt_row, column=tt_col)
            # İstersen sabit yaz:
            # tt_cell.value = 60
            tt_cell.value = None
            tt_cell.alignment = self.center



    
        """
        ORTA bloktaki (Westinghouse) SOP tablosunu doldurur.
        Layout, Zaman Etüdü SOP ile aynı:
          No | ID | OPERASYON ADIMLARI | ÖNCÜL | VA | NVAN | NVA
        Süre kaynağı: west_map (Westinghouse için kullandığımız süre haritası).
        """

        # Bu sefer SOL değil, ORTA blok → başlangıç sütunu = 11
        sc = 11                   # Westinghouse SOP bandının başlangıç sütunu
        sr = start_row            # SOP'un başladığı satır (bizde 2)
        ec = sc + 8

        # _write_sop_band içindeki aynı satır hesapları:
        h = sr + 11               # tablo başlık satırı (No, ID, ...)
        n = max(10, adim_sayisi)  # en az 10 satır
        data_start = h + 1        # ilk operasyon satırı
        data_end = h + n          # son operasyon satırı
        toplam_row = h + 1 + n    # "TOPLAM SÜRE" satırı

        # İş adımlarını hazırlayalım
        steps = job_df.dropna(subset=["AdımID"]).copy()
        if steps.empty:
            return

        steps["AdımID"] = steps["AdımID"].astype(int)
        steps = steps.sort_values("AdımID")

        # DegerTuru -> VA / NVAN / NVA sınıflaması (Zaman Etüdü ile aynı mantık)
        def _classify(deger):
            if not isinstance(deger, str):
                return ""
            d = deger.strip().upper()
            if d in ("KD", "VA"):
                return "VA"
            if d in ("ZGK", "NVAN", "NVAA", "NV-AN"):
                return "NVAN"
            if d in ("KDZ", "NVA"):
                return "NVA"
            return ""

        current_row = data_start
        no = 1

        for _, r in steps.iterrows():
            if current_row > data_end:
                break

            adim_id = r.get("AdımID")
            op_name = r.get("Adım Adı")
            oncul = r.get("Öncül Adım")
            cat = _classify(r.get("DegerTuru"))

            # Westinghouse için süre kaynağı: west_map
            sure = float(west_map.get(adim_id, 0.0) or 0.0)

            va = nv_an = nva = 0.0
            if cat == "VA":
                va = sure
            elif cat == "NVAN":
                nv_an = sure
            elif cat == "NVA":
                nva = sure

            # Hücrelere yaz (ORTA blok, sc=11'den başlıyor)
            ws.cell(row=current_row, column=sc,     value=no).alignment = self.center    # No
            ws.cell(row=current_row, column=sc + 1, value=adim_id).alignment = self.center
            ws.cell(row=current_row, column=sc + 2, value=op_name).alignment = self.left
            ws.cell(row=current_row, column=sc + 4, value=oncul).alignment = self.center
            ws.cell(row=current_row, column=sc + 5, value=round(va, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 6, value=round(nv_an, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 7, value=round(nva, 4)).alignment = self.center

            # Kenarlık
            for col in range(sc, sc + 8):
                ws.cell(row=current_row, column=col).border = self.thin_border

            no += 1
            current_row += 1

        # --- TOPLAM SÜRE satırı için formüller (VA / NVAN / NVA) ---
        if data_start <= data_end:
            # VA toplamı
            first_va = ws.cell(row=data_start, column=sc + 5).coordinate
            last_va  = ws.cell(row=data_end,   column=sc + 5).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 5,
                value=f"=SUM({first_va}:{last_va})"
            ).alignment = self.center

            # NVAN toplamı
            first_nv = ws.cell(row=data_start, column=sc + 6).coordinate
            last_nv  = ws.cell(row=data_end,   column=sc + 6).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 6,
                value=f"=SUM({first_nv}:{last_nv})"
            ).alignment = self.center

            # NVA toplamı
            first_nva = ws.cell(row=data_start, column=sc + 7).coordinate
            last_nva  = ws.cell(row=data_end,   column=sc + 7).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 7,
                value=f"=SUM({first_nva}:{last_nva})"
            ).alignment = self.center

    def _fill_most_sop_from_most(
        self,
        ws,
        start_row: int,
        adim_sayisi: int,
        job_df: pd.DataFrame,
        most_map: Dict
    ) -> None:
        """
        Sağdaki (MOST) SOP tablosunu doldurur.
        Kaynaklar:
          - Is_Adimlari.csv  -> AdımID, Adım Adı, Öncül Adım, DegerTuru
          - Basic/Mini/Maxi_Most_Analizleri.csv -> AdımID bazında ToplamSaniye (most_map)
        """

        # MOST SOP bandı sağ blok: start_col = 21
        sc = 21                  # başlangıç sütunu (MOST)
        sr = start_row           # SOP'un başladığı satır (bizde 2)
        ec = sc + 8

        # _write_sop_band içindeki aynı hesaplar:
        h = sr + 11              # tablo başlık satırı (No, ID, ...)
        n = max(10, adim_sayisi) # en az 10 satır
        data_start = h + 1       # ilk operasyon satırı
        data_end = h + n         # son operasyon satırı
        toplam_row = h + 1 + n   # "TOPLAM SÜRE" satırı

        # İş adımlarını hazırlayalım
        steps = job_df.dropna(subset=["AdımID"]).copy()
        if steps.empty:
            return

        # AdımID int olsun ve sıralı gidelim
        steps["AdımID"] = steps["AdımID"].astype(int)
        steps = steps.sort_values("AdımID")

        # DegerTuru -> VA / NVAN / NVA sınıflaması (diğer SOP'larla aynı)
        def _classify(deger):
            if not isinstance(deger, str):
                return ""
            d = deger.strip().upper()
            if d in ("KD", "VA"):
                return "VA"
            if d in ("ZGK", "NVAN", "NVAA", "NV-AN"):
                return "NVAN"
            if d in ("KDZ", "NVA"):
                return "NVA"
            return ""

        no = 1
        current_row = data_start

        for _, row in steps.iterrows():
            if current_row > data_end:
                break

            adim_id = int(row["AdımID"])
            op_name = row.get("Adım Adı", "")
            oncul   = row.get("Öncül Adım", "")

            cat = _classify(row.get("DegerTuru", ""))

            # MOST süresi (sn) – ToplamSaniye haritasından
            sure = float(most_map.get(adim_id, 0.0) or 0.0)

            # DegerTuru'na göre VA / NVAN / NVA dağıt
            va = nv_an = nva = 0.0
            if cat == "VA":
                va = sure
            elif cat == "NVAN":
                nv_an = sure
            elif cat == "NVA":
                nva = sure

            # Hücrelere yaz (sağdaki SOP bandı)
            ws.cell(row=current_row, column=sc,     value=no).alignment = self.center    # No
            ws.cell(row=current_row, column=sc + 1, value=adim_id).alignment = self.center
            ws.cell(row=current_row, column=sc + 2, value=op_name).alignment = self.left
            ws.cell(row=current_row, column=sc + 4, value=oncul).alignment = self.center
            ws.cell(row=current_row, column=sc + 5, value=round(va, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 6, value=round(nv_an, 4)).alignment = self.center
            ws.cell(row=current_row, column=sc + 7, value=round(nva, 4)).alignment = self.center

            # Kenarlık
            for col in range(sc, sc + 8):
                ws.cell(row=current_row, column=col).border = self.thin_border

            no += 1
            current_row += 1

        # --- TOPLAM SÜRE satırı için formüller (VA / NVAN / NVA) ---
        if data_start <= data_end:
            # VA toplamı
            first_va = ws.cell(row=data_start, column=sc + 5).coordinate
            last_va  = ws.cell(row=data_end,   column=sc + 5).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 5,
                value=f"=SUM({first_va}:{last_va})"
            ).alignment = self.center

            # NVAN toplamı
            first_nv = ws.cell(row=data_start, column=sc + 6).coordinate
            last_nv  = ws.cell(row=data_end,   column=sc + 6).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 6,
                value=f"=SUM({first_nv}:{last_nv})"
            ).alignment = self.center

            # NVA toplamı
            first_nva = ws.cell(row=data_start, column=sc + 7).coordinate
            last_nva  = ws.cell(row=data_end,   column=sc + 7).coordinate
            ws.cell(
                row=toplam_row,
                column=sc + 7,
                value=f"=SUM({first_nva}:{last_nva})"
            ).alignment = self.center





    def _write_sop_triplet(self, ws, start_row, adim_sayisi):
        end1 = self._write_sop_band(ws, start_row, 1,  "Zaman Etüdü",  adim_sayisi)
        end2 = self._write_sop_band(ws, start_row, 11, "Westinghouse", adim_sayisi)
        end3 = self._write_sop_band(ws, start_row, 21, "MOST",        adim_sayisi)
        return max(end1, end2, end3)



    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        if name is None:
            name = "Is_Adi"
        name = re.sub(r'[\\/*?:\[\]]', '_', str(name))
        return name[:31]
    
    def _write_main_sheet(self, wb: Workbook):
        """
        Ana Sayfa:
        İstasyon | Operasyon Adı | VA (sn) | NVAN (sn) | NVA (sn) | Ortalama Süre (sn) | Analiz Zamanı
        """
        ws = wb.create_sheet("Ana Sayfa", 0)  # En başa ekle

        headers = [
            "İstasyon",
            "Operasyon Adı",
            "VA (sn)",
            "NVAN (sn)",
            "NVA (sn)",
            "Ortalama Süre (sn)",
            "Analiz Zamanı",
        ]

                # Başlık satırı
        from openpyxl.styles import PatternFill  # dosyanın başında zaten import varsa tekrar gerek yok

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = self.header_font

            # VA / NVAN / NVA için özel renkler
            if h == "VA (sn)":
                c.fill = PatternFill("solid", fgColor="00B050")  # yeşil
            elif h == "NVAN (sn)":
                c.fill = PatternFill("solid", fgColor="FFFF00")  # sarı
            elif h == "NVA (sn)":
                c.fill = PatternFill("solid", fgColor="FF0000")  # kırmızı
            else:
                c.fill = self.header_fill

            c.alignment = self.center
            c.border = self.thin_border


        # Zaman etüdü: JobID–AdımID bazında ortalama süre
        step_avg = pd.DataFrame(columns=["JobID", "AdımID", "Sure"])
        if hasattr(self, "zaman_etudu") and not self.zaman_etudu.empty:
            zdf = self.zaman_etudu.copy()
            if "JobID" in zdf.columns and "AdımID" in zdf.columns:
                step_avg = (
                    zdf.groupby(["JobID", "AdımID"])["Sure"]
                    .mean()
                    .reset_index()
                )

        current_row = 2

        # Her İş (Operasyon) için bir satır
        for job_name, job_df in self.is_adimlari.groupby("İş Adı"):
            # JobID
            job_id = None
            if "JobID" in job_df.columns:
                s = job_df["JobID"].dropna()
                if not s.empty:
                    job_id = int(s.iloc[0])

            # Bu işe ait adımlar
            steps = job_df.dropna(subset=["AdımID"]).copy()
            if steps.empty:
                continue

            # İstasyon = en sık geçen İş İstasyonu
            istasyon = ""
            if "İş İstasyonu" in steps.columns:
                ist_ser = steps["İş İstasyonu"].dropna()
                if not ist_ser.empty:
                    istasyon = ist_ser.value_counts().idxmax()

            # Adım ortalamalarını bu iş ile birleştir
            merged = steps.copy()
            merged["Sure"] = 0.0

            if job_id is not None and not step_avg.empty:
                m = merged.merge(
                    step_avg,
                    on=["JobID", "AdımID"],
                    how="left",
                    suffixes=("", "_avg"),
                )
                merged["Sure"] = m["Sure_avg"].fillna(0.0)
            elif not step_avg.empty:
                # JobID kullanılamıyorsa sadece AdımID üzerinden dener
                m = merged.merge(
                    step_avg[["AdımID", "Sure"]],
                    on="AdımID",
                    how="left",
                )
                merged["Sure"] = m["Sure"].fillna(0.0)

            # VA / NVAN / NVA toplamları
            va = nv_an = nva = 0.0

            def _classify(deger_turu: str):
                dt = str(deger_turu or "").strip().upper()
                if dt in ("KD", "VA"):
                    return "VA"
                if dt in ("ZGK", "KDG", "NVAN"):
                    return "NVAN"
                if dt in ("KDZ", "KDS", "NVA"):
                    return "NVA"
                return None

            for _, r in merged.iterrows():
                sure = float(r.get("Sure") or 0.0)
                cat = _classify(r.get("DegerTuru"))
                if cat == "VA":
                    va += sure
                elif cat == "NVAN":
                    nv_an += sure
                elif cat == "NVA":
                    nva += sure

            toplam = va + nv_an + nva

            # Analiz Zamanı: Zaman Etüdü veya Westinghouse içindeki en güncel Timestamp
            analiz_zamani = ""
            timestamps = []

            if job_id is not None and hasattr(self, "zaman_etudu") and not self.zaman_etudu.empty:
                ts = self.zaman_etudu[self.zaman_etudu["JobID"] == job_id]["Timestamp"].dropna()
                timestamps.extend(ts.tolist())

            if job_id is not None and not self.west.empty and "Timestamp" in self.west.columns:
                ts = self.west[self.west["JobID"] == job_id]["Timestamp"].dropna()
                timestamps.extend(ts.tolist())

            if timestamps:
                # ISO tarih formatı olduğu için string max en yeni zamana denk gelir
                analiz_zamani = max(timestamps)

            row_vals = [
                istasyon,
                job_name,
                round(va, 4) if va else 0,
                round(nv_an, 4) if nv_an else 0,
                round(nva, 4) if nva else 0,
                round(toplam, 4) if toplam else 0,
                analiz_zamani,
            ]

            for col, val in enumerate(row_vals, start=1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.border = self.thin_border
                if col in (3, 4, 5, 6):
                    c.alignment = self.center
                else:
                    c.alignment = self.left

            # Operasyon Adı hücresine hyperlink ekle (ilgili iş sayfasına gider)
                    # Operasyon Adı hücresine hyperlink ekle (ilgili iş sayfasına gider)
        # job_to_sheet'ten gerçek sheet adını al, yoksa safe name'e düş
            sheet_name = self.job_to_sheet.get(job_name, self._safe_sheet_name(job_name))

            op_cell = ws.cell(row=current_row, column=2)
            op_cell.value = job_name  # zaten row_vals içinde yazıldı ama garanti olsun

            # Excel içi hyperlink: location = "'Sayfa Adı'!A1" formatında olmalı
            location = f"'{sheet_name}'!A1"
            op_cell.hyperlink = Hyperlink(ref=op_cell.coordinate, location=location, display=job_name)
            op_cell.style = "Hyperlink"


            current_row += 1

        self._autosize(ws)


    # ------------ Westinghouse Bloğu (SOL) ------------

    def _write_west_block(self, ws, start_row: int, step_name: str, wh_row: pd.Series) -> int:
        """
        Westinghouse bloğunu A..G sütunlarında yazar.
        A: Adım Adı (13 satır merge)
        B..G: başlıklar ve içerik
        Geri dönüş: bir sonraki blok için başlangıç satırı (araya 1 boş satırla)
        """
        col_start = 1  # A sütunu

        # Başlık satırı (blok üst satırı): B..G
        headers = self.WH_HEADERS

        # Adım Adı hücresi: 13 satır birleşik (A sütunu)
        ws.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row + 12, end_column=col_start)
        c_adim = ws.cell(row=start_row, column=col_start)
        c_adim.value = step_name
        c_adim.alignment = self.center
        c_adim.border = self.thin_border

        # B..G başlıkları
        for i, h in enumerate(headers, start=col_start + 1):
            cell = ws.cell(row=start_row, column=i)
            cell.value = h
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center
            cell.border = self.thin_border

        # Performans Faktörleri (4 satır) – başlığın altından itibaren
        for i, nm in enumerate(self.WH_PF_NAMES):
            r = start_row + 1 + i
            ws.cell(row=r, column=col_start + 1, value=nm).alignment = self.left
            ws.cell(row=r, column=col_start + 1).border = self.thin_border
            ws.cell(row=r, column=col_start + 2, value=str(wh_row.get(nm, "-")) if pd.notna(wh_row.get(nm, None)) else "-").alignment = self.left
            ws.cell(row=r, column=col_start + 2).border = self.thin_border

        # Zorluk Faktörleri (7 satır) – başlığın altından itibaren
        zf_map = {
            "Kişisel Gereksinimler": str(wh_row.get("Kisisel_Gereksinimler", "-")) if pd.notna(wh_row.get("Kisisel_Gereksinimler", None)) else "-",
            "Fiziksel Çaba": _pick_label(wh_row, [
                ("Fiziksel_Caba_Cok_Hafif", "Çok Hafif"),
                ("Fiziksel_Caba_Hafif", "Hafif"),
                ("Fiziksel_Caba_Orta", "Orta"),
                ("Fiziksel_Caba_Agir", "Ağır"),
                ("Fiziksel_Caba_Cok_Agir", "Çok Ağır"),
            ]),
            "Düşünsel Çaba": _pick_label(wh_row, [
                ("Dusunsel_Caba_Plan_Normal", "Plan/Normal"),
                ("Dusunsel_Caba_Karisik_Normal", "Karışık/Normal"),
                ("Dusunsel_Caba_Plan_Yogun", "Plan/Yoğun"),
                ("Dusunsel_Caba_Karisik_Yogun", "Karışık/Yoğun"),
            ]),
            "Çalışma Pozisyonu": _pick_label(wh_row, [
                ("Poz_Serbest", "Serbest"),
                ("Poz_Sabit_Durus", "Sabit Duruş"),
                ("Poz_Sabit_Ayakta", "Sabit Ayakta"),
                ("Poz_Cokme_Egilme", "Çökme/Eğilme"),
                ("Poz_Uzanma_ve_Omuz", "Uzanma/Omuz"),
            ]),
            "Atmosfer": _pick_label(wh_row, [
                ("Atmosfer_Temiz", "Temiz"),
                ("Atmosfer_Kotu_Koku", "Kötü Koku"),
                ("Atmosfer_Zararlı_Toz_Gaz", "Zararlı Toz/Gaz"),
            ]),
            "Isı": _pick_label(wh_row, [
                ("Isı_Soguk", "Soğuk"),
                ("Isı_Normal", "Normal"),
                ("Isı_Sicak", "Sıcak"),
            ]),
            "Gürültü": _pick_label(wh_row, [
                ("Gurultu_Normal_Is", "Normal (İş)"),
                ("Gurultu_Normal_Makine", "Normal (Makine)"),
                ("Gurultu_Yuksek_Sabit", "Yüksek (Sabit)"),
                ("Gurultu_Yuksek_Frekans", "Yüksek (Frekans)"),
            ]),
        }

        base = start_row + 1
        for i, nm in enumerate(self.WH_ZF_NAMES):
            r = base + i
            ws.cell(row=r, column=col_start + 3, value=nm).alignment = self.left
            ws.cell(row=r, column=col_start + 3).border = self.thin_border
            ws.cell(row=r, column=col_start + 4, value=zf_map.get(nm, "-")).alignment = self.left
            ws.cell(row=r, column=col_start + 4).border = self.thin_border

        # Genel Koşullar / Koruyucu Ekipmanlar – başlığın altından itibaren
        gk_list = _multi_labels(wh_row, [
            ("Genel_Kirli", "Kirli"),
            ("Genel_Islak_Doseme", "Islak Döşeme"),
            ("Genel_Titresim", "Titreşim"),
            ("Genel_Monotonluk", "Monotonluk"),
            ("Genel_Dusunsel_Yorgunluk", "Düşünsel Yorgunluk"),
        ])
        ke_list = _multi_labels(wh_row, [
            ("Koruyucu_Elbise_Takım", "Takım"),
            ("Koruyucu_Elbise_Eldiven", "Eldiven"),
            ("Koruyucu_Elbise_Agır_ve_Ozel_Yelek", "Ağır/Özel Yelek"),
            ("Koruyucu_Elbise_Maske", "Maske"),
        ])

        # Başlığın altındaki ilk hücreden başlayarak her birini alt alta yaz
        gk_start = start_row + 1
        for i in range(12):  # 12 satıra kadar yaz (PF+ZF toplam satır sayısıyla hizalı görünür)
            val = gk_list[i] if i < len(gk_list) else ("" if len(gk_list) > 0 else ("-" if i == 0 else ""))
            c = ws.cell(row=gk_start + i, column=col_start + 5)
            c.value = val
            c.alignment = self.left
            c.border = self.thin_border

        ke_start = start_row + 1
        for i in range(12):
            val = ke_list[i] if i < len(ke_list) else ("" if len(ke_list) > 0 else ("-" if i == 0 else ""))
            c = ws.cell(row=ke_start + i, column=col_start + 6)
            c.value = val
            c.alignment = self.left
            c.border = self.thin_border

        # Tüm blok sınırları
        for rr in range(start_row, start_row + 13):
            for cc in range(col_start, col_start + 7):
                ws.cell(row=rr, column=cc).border = self.thin_border

        return start_row + 13 + 1  # 1 satır boşlukla bir sonraki blok

    # ------------ MOST (SAĞ) ------------

        # ------------ MOST (SAĞ) ------------

    def _write_most_table(self, ws, start_row: int, merged: pd.DataFrame) -> None:
        """MOST tablosunu J..P aralığında yazar."""
        col_start = 10  # J

        header_row = start_row
        # Başlık satırı
        for i, h in enumerate(self.MOST_HEADERS, start=col_start):
            cell = ws.cell(row=header_row, column=i)
            cell.value = h
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center
            cell.border = self.thin_border

        # Genel tekrar haritası
        genel_tekrar_map = (
            merged[merged["ParametreKodu"].apply(_is_genel_tekrar)]
            .groupby("AnalizID")["TekrarSayisi"]
            .max()
            .to_dict()
        )

        # İçerik
        out_rows: List[List] = []
        group_sizes: List[int] = []

        for analiz_id, grp in merged.groupby("AnalizID", sort=False):
            grp_eff = grp[~grp["ParametreKodu"].apply(_is_genel_tekrar)].copy()
            gsize = len(grp_eff) if len(grp_eff) > 0 else 1
            group_sizes.append(gsize)

            genel_tekrar = int(genel_tekrar_map.get(analiz_id, 1))
            first = grp_eff.iloc[0] if not grp_eff.empty else grp.iloc[0]
            step_name = first.get("Adım Adı", "") or ""
            model_left, sekans = _split_model_and_sekans(first.get("ModelTipi", "") or "")

            if grp_eff.empty:
                out_rows.append([step_name, model_left, sekans, genel_tekrar, "", "", ""])
            else:
                for _, r in grp_eff.iterrows():
                    indis = _extract_indis(r.get("ParametreKodu", ""))
                    indis_tekrari = r.get("TekrarSayisi", "")
                    aciklama = r.get("SecilenDeger", "")
                    out_rows.append([step_name, model_left, sekans, genel_tekrar, indis, indis_tekrari, aciklama])

        # Yaz ve birleştir (AnalizID bloğu kadar)
        pointer = header_row + 1
        idx = 0
        for gsize in group_sizes:
            # satırları yaz
            for row in out_rows[idx: idx + gsize]:
                for j, v in enumerate(row, start=col_start):
                    c = ws.cell(row=pointer, column=j)
                    c.value = v
                    c.border = self.thin_border
                    c.alignment = self.center if j - col_start + 1 in (1, 2, 3, 4, 5) else self.left
                pointer += 1

            # birleştirmeler
            if gsize > 1:
                for off in range(0, 4):  # Adım Adı, Model, Sekans, Genel Tekrar
                    ws.merge_cells(start_row=pointer - gsize, start_column=col_start + off,
                                   end_row=pointer - 1, end_column=col_start + off)
                    ws.cell(row=pointer - gsize, column=col_start + off).alignment = self.center

            idx += gsize

        # Otomatik genişlik
        for col_idx in range(col_start, col_start + len(self.MOST_HEADERS)):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                val = "" if val is None else str(val)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # ------------ CPM + Gantt Yardımcıları ------------

    def _build_cpm_rows_linear(self, steps_df: pd.DataFrame, durations_map: Dict) -> List[Dict]:
        """
        CPM tablosunu hesaplar ve 'DegerTuru' bilgisini de içerir.
        Sonuçları Erken Başlama (ES) süresine göre sıralar (Gantt şelalesi için).
        """
        from collections import deque

        steps = steps_df.dropna(subset=["AdımID"]).copy()
        if steps.empty:
            return []

        steps["AdımID"] = steps["AdımID"].astype(int)
        ids = steps["AdımID"].tolist()

        # Süre haritası
        duration = {aid: float(durations_map.get(aid, 0.0) or 0.0) for aid in ids}

        # Öncelik ve ardıllar
        preds: Dict[int, List[int]] = {}
        succs: Dict[int, List[int]] = {aid: [] for aid in ids}
        indegree: Dict[int, int] = {aid: 0 for aid in ids}

        for _, row in steps.iterrows():
            aid = int(row["AdımID"])
            raw_pred = row.get("Öncül Adım", "")
            plist = _parse_predecessors(raw_pred)
            plist = [p for p in plist if p in ids]
            preds[aid] = plist

        for aid in ids:
            for p in preds.get(aid, []):
                succs.setdefault(p, []).append(aid)
                indegree[aid] += 1

        # Topolojik sıralama
        q = deque([aid for aid in ids if indegree[aid] == 0])
        topo: List[int] = []
        while q:
            n = q.popleft()
            topo.append(n)
            for s in succs.get(n, []):
                indegree[s] -= 1
                if indegree[s] == 0:
                    q.append(s)

        for aid in ids:
            if aid not in topo: topo.append(aid)

        # İleri geçiş (ES / EF)
        ES: Dict[int, float] = {aid: 0.0 for aid in ids}
        EF: Dict[int, float] = {}

        for aid in topo:
            pl = preds.get(aid, [])
            if pl:
                ES[aid] = max(EF[p] for p in pl if p in EF)
            else:
                ES.setdefault(aid, 0.0)
            EF[aid] = ES[aid] + duration[aid]

        project_duration = max(EF.values()) if EF else 0.0

        # Geri geçiş (LS / LF)
        LS: Dict[int, float] = {}
        LF: Dict[int, float] = {}
        
        last_activities = [aid for aid in ids if not succs.get(aid)]
        for aid in last_activities:
            LF[aid] = project_duration
            LS[aid] = project_duration - duration[aid]

        for aid in reversed(topo):
            if aid in LS: continue
            succ_list = succs.get(aid, [])
            if succ_list:
                ls_candidates = [LS[s] for s in succ_list if s in LS]
                lf_val = min(ls_candidates) if ls_candidates else project_duration
            else:
                lf_val = project_duration
            LF[aid] = lf_val
            LS[aid] = lf_val - duration[aid]

        rows: List[Dict] = []
        for _, row in steps.iterrows():
            aid = int(row["AdımID"])
            
            # --- DÜZELTME BURADA YAPILDI ---
            # Değerleri önce değişkenlere alıp, undefined hatasını önlüyoruz.
            current_es = ES.get(aid, 0.0)
            current_ef = EF.get(aid, current_es + duration[aid])
            current_ls = LS.get(aid, current_es)
            current_lf = LF.get(aid, current_ef)
            current_slack = current_ls - current_es
            
            rows.append({
                "AdımID": aid,
                "Adım Adı": row.get("Adım Adı", "") or "",
                "Öncül": ",".join(str(p) for p in preds.get(aid, [])),
                "Süre": duration[aid],
                "ES": current_es,
                "EF": current_ef,
                "LS": current_ls,
                "LF": current_lf,
                "Bolluk": current_slack,
                "DegerTuru": str(row.get("DegerTuru", "")).strip().upper()
            })

        # Listeyi Erken Başlama (ES) zamanına göre sırala -> Gantt'ta şelale görünümü için
        rows.sort(key=lambda x: (x["ES"], x["AdımID"]))
        
        return rows

    def _write_cpm_and_gantt_block(self, ws, start_row: int, col_start: int,
                                   title: str, cpm_rows: List[Dict]) -> int:
        """
        Üstte CPM verileri, altta GANTT Şeması oluşturur.
        Gantt şeması: 
        - İlk iş en üstte (şelale yapısı).
        - Sadece süre çubukları görünür (başlangıç kısmı şeffaf).
        - Çubuklar VA (Yeşil), NVAN (Sarı), NVA (Kırmızı) olarak renklendirilir.
        """
        if not cpm_rows:
            return start_row

        # --- 1. CPM VERİ TABLOSU ---
        headers = ["AdımID", "Adım Adı", "Öncül Adım", "Süre (sn)", "ES", "EF", "LS", "LF", "Bolluk"]
        end_col = col_start + len(headers) - 1

        title_cell = ws.cell(row=start_row, column=col_start, value=title)
        ws.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row, end_column=end_col)
        title_cell.font = self.header_font
        title_cell.alignment = self.center

        header_row = start_row + 1
        for off, h in enumerate(headers):
            c = ws.cell(row=header_row, column=col_start + off, value=h)
            c.font = self.header_font
            c.fill = self.header_fill
            c.alignment = self.center
            c.border = self.thin_border

        data_start = header_row + 1
        for idx, row in enumerate(cpm_rows):
            rr = data_start + idx
            vals = [
                row["AdımID"], row["Adım Adı"], row["Öncül"], row["Süre"],
                row["ES"], row["EF"], row["LS"], row["LF"], row["Bolluk"]
            ]
            for off, v in enumerate(vals):
                cc = ws.cell(row=rr, column=col_start + off, value=v)
                cc.border = self.thin_border
                if off in (0, 3, 4, 5, 6, 7, 8):
                    cc.alignment = self.center
                else:
                    cc.alignment = self.left

        last_cpm_row = data_start + len(cpm_rows) - 1

        # --- 2. GANTT VERİ HAZIRLIĞI (Stacked Bar Mantığı) ---
        # Excel'de Gantt yapmak için:
        # Seri 1: Başlangıç (ES) -> Şeffaf/Görünmez yapılır (Barı sağa iter)
        # Seri 2: VA Süresi -> Yeşil
        # Seri 3: NVAN Süresi -> Sarı
        # Seri 4: NVA Süresi -> Kırmızı
        
        gantt_data_start_row = last_cpm_row + 3
        
        # Başlıklar (Grafik kaynağı olacak gizli tablo)
        g_headers = ["Adım Adı", "Başlangıç (ES)", "VA", "NVAN", "NVA"]
        for i, h in enumerate(g_headers):
            ws.cell(row=gantt_data_start_row, column=col_start + i, value=h)

        # Verileri ayrıştırarak yaz
        for idx, row in enumerate(cpm_rows):
            rr = gantt_data_start_row + 1 + idx
            
            # Adım Adı (Y Ekseni Etiketi)
            ws.cell(row=rr, column=col_start, value=row["Adım Adı"])
            
            # Başlangıç Zamanı (Görünmez Olacak Seri)
            ws.cell(row=rr, column=col_start + 1, value=row["ES"])
            
            # Süreyi Değer Türüne Göre İlgili Sütuna Yaz, Diğerlerine 0 Yaz
            sure = row["Süre"]
            d_turu = row["DegerTuru"]
            
            va_val = sure if d_turu in ("VA", "KD") else 0
            nvan_val = sure if d_turu in ("NVAN", "ZGK", "NVAA") else 0
            nva_val = sure if d_turu in ("NVA", "KDZ", "KDS") else 0
            
            # Eğer tür belirtilmemişse varsayılan olarak NVA kabul edelim (ya da tercihe göre VA)
            if va_val == 0 and nvan_val == 0 and nva_val == 0:
                nva_val = sure

            ws.cell(row=rr, column=col_start + 2, value=va_val)   # VA (Yeşil)
            ws.cell(row=rr, column=col_start + 3, value=nvan_val) # NVAN (Sarı)
            ws.cell(row=rr, column=col_start + 4, value=nva_val)  # NVA (Kırmızı)

        # --- 3. GANTT GRAFİĞİNİ OLUŞTURMA ---
        chart = BarChart()
        chart.type = "bar"      # Yatay Çubuk
        chart.grouping = "stacked" # Yığılmış
        chart.overlap = 100
        chart.title = title.replace("CPM + Gantt", "Gantt Şeması")
        chart.style = 10 # Genel stil
        
        # --- Seri Tanımları ---
        data_rows_count = len(cpm_rows)
        
        # X Ekseni (Kategoriler: Adım Adları)
        cats = Reference(ws, min_col=col_start, min_row=gantt_data_start_row+1, max_row=gantt_data_start_row+data_rows_count)
        chart.set_categories(cats)

        # 1. Seri: Başlangıç (ES) -> GÖRÜNMEZ YAP (No Fill)
        es_data = Reference(ws, min_col=col_start+1, min_row=gantt_data_start_row, max_row=gantt_data_start_row+data_rows_count)
        s_es = Series(es_data, title_from_data=True)
        s_es.graphicalProperties.noFill = True # Dolgu yok, sadece öteleme sağlar
        chart.series.append(s_es)

        # 2. Seri: VA (Yeşil)
        va_data = Reference(ws, min_col=col_start+2, min_row=gantt_data_start_row, max_row=gantt_data_start_row+data_rows_count)
        s_va = Series(va_data, title_from_data=True)
        s_va.graphicalProperties.solidFill = "00B050" # Yeşil
        s_va.graphicalProperties.line.noFill = True
        chart.series.append(s_va)

        # 3. Seri: NVAN (Sarı)
        nvan_data = Reference(ws, min_col=col_start+3, min_row=gantt_data_start_row, max_row=gantt_data_start_row+data_rows_count)
        s_nvan = Series(nvan_data, title_from_data=True)
        s_nvan.graphicalProperties.solidFill = "FFFF00" # Sarı
        s_nvan.graphicalProperties.line.noFill = True
        chart.series.append(s_nvan)

        # 4. Seri: NVA (Kırmızı)
        nva_data = Reference(ws, min_col=col_start+4, min_row=gantt_data_start_row, max_row=gantt_data_start_row+data_rows_count)
        s_nva = Series(nva_data, title_from_data=True)
        s_nva.graphicalProperties.solidFill = "FF0000" # Kırmızı
        s_nva.graphicalProperties.line.noFill = True
        chart.series.append(s_nva)

        # --- Grafik Ayarları ---
        # Y Ekseni (İşler): Ters sırala ki ilk iş en üstte olsun
        chart.y_axis.scaling.orientation = "maxMin" 
        chart.y_axis.majorTickMark = "out"
        
        # X Ekseni (Zaman): Izgara çizgileri
        from openpyxl.chart.axis import ChartLines
        chart.x_axis.majorGridlines = ChartLines() # Dikey zaman çizgileri
        chart.x_axis.title = "Zaman (sn)"
        
        # Boyut ve Konum
        chart.height = 10 + (data_rows_count * 0.5) # İş sayısına göre yükseklik
        chart.width = 18
        
        anchor_cell = f"{get_column_letter(col_start)}{gantt_data_start_row + data_rows_count + 2}"
        ws.add_chart(chart, anchor_cell)

        return gantt_data_start_row + data_rows_count + 20 # Sonraki içerik için boşluk



    def _write_sheet_for_job(self, wb: Workbook, job_name: str, job_df: pd.DataFrame):
        ws = wb.create_sheet(self._safe_sheet_name(job_name))
        
        # Bu işe ait adımlar
        adim_ids = job_df["AdımID"].dropna().unique().tolist()
        if not adim_ids:
            return

        # Adım listesi (CPM için) – Öncül Adım ve DegerTuru kolonlarını al
        steps_df = (
            job_df[["AdımID", "Adım Adı", "Öncül Adım", "DegerTuru"]]
            .dropna(subset=["AdımID"])
            .drop_duplicates()
        )


        # Bu işe ait MOST analizleri
        ilgili_analizler = self.analiz_all[self.analiz_all["AdımID"].isin(adim_ids)].copy()

        # MOST detaylarını birleştir (mevcut MOST tablosu için)
        merged = self.detay_all.merge(ilgili_analizler, on="AnalizID", how="right")
        merged = merged.merge(job_df[["AdımID", "Adım Adı"]].drop_duplicates(),
                              on="AdımID", how="left")

        # -------- CPM için süre haritaları --------
        # Zaman Etüdü süreleri
        zaman_map: Dict = {}
        if hasattr(self, "zaman_etudu") and not self.zaman_etudu.empty:
            job_id = None
            if "JobID" in job_df.columns:
                s = job_df["JobID"].dropna()
                if not s.empty:
                    job_id = s.iloc[0]

            zdf = self.zaman_etudu.copy()
            if job_id is not None and "JobID" in zdf.columns:
                zdf = zdf[zdf["JobID"] == job_id]
            if not zdf.empty:
                zaman_map = zdf.groupby("AdımID")["Sure"].mean().to_dict()

        # MOST süreleri (ToplamSaniye)
        df_most = (
            self.analiz_all
            .dropna(subset=["ToplamSaniye"])
            .copy()
        )

        if "AnalizID" in df_most.columns:
            # AnalizID varsa: en büyük AnalizID = en son yapılan analiz
            df_most = df_most.sort_values(["AdımID", "AnalizID"])
            last_rows = df_most.groupby("AdımID").tail(1)
        else:
            # AnalizID yoksa: aynı AdımID'lerde en son satırı kabul et
            df_most = df_most.sort_values(["AdımID"])
            last_rows = df_most.groupby("AdımID").tail(1)

        most_map = (
            last_rows
            .set_index("AdımID")["ToplamSaniye"]
            .astype(float)
            .to_dict()
        )


       
       # -------- CPM için süre haritaları --------

        # 1) Zaman Etüdü haritası (mevcut kodun kalsın)

        # 2) MOST haritası (mevcut kodun kalsın)

        # 3) Westinghouse haritası – ÖNCE StandartZaman, yoksa eskisi gibi fallback
        west_map: Dict[int, float] = {}

        if not self.west.empty and "StandartZaman" in self.west.columns:
            # Bu işe ait adımlar:
            adim_ids = steps_df["AdımID"].dropna().astype(int).unique().tolist()
            wdf = self.west[self.west["AdımID"].isin(adim_ids)].copy()
            if not wdf.empty:
                # Her adım için en son yapılan analizi al
                wdf = wdf.sort_values(["AdımID", "AnalizID"])
                last = wdf.groupby("AdımID").tail(1)
                west_map = last.set_index("AdımID")["StandartZaman"].dropna().astype(float).to_dict()

        # Eğer hâlâ boşsa, eskisi gibi Zaman Etüdü / MOST fallback kullan
        if not west_map:
            if zaman_map:
                west_map = zaman_map
            elif most_map:
                west_map = most_map


       
        

        # -------- Zaman Etüdü ve MOST için fallback mantığı --------

        # Zaman Etüdü haritası boşsa, en azından MOST'tan doldur
        if not zaman_map and most_map:
            zaman_map = most_map

        # Hâlâ boşsa, tüm adımlar için 0 ver (grafikler bozulmasın)
        if not zaman_map:
            zaman_map = {aid: 0.0 for aid in steps_df["AdımID"]}

        # MOST boşsa, Zaman Etüdü'ne yaslansın
        if not most_map:
            most_map = zaman_map


        # -------- Üstte 3 CPM + 3 Gantt (yan yana) --------
               # Bu iş için toplam adım sayısı (SOP satır sayısını belirleyecek)
        adim_sayisi = len(steps_df)

        # -------- Üstte 3 SOP + altında 3 CPM + 3 Gantt --------
        sop_bottom = self._write_sop_triplet(ws, start_row=2, adim_sayisi=adim_sayisi)


               # Zaman Etüdü SOP tablosu
        self._fill_zaman_sop_from_time_study(
            ws=ws,
            start_row=2,
            adim_sayisi=adim_sayisi,
            job_df=job_df,
            zaman_map=zaman_map
        )

               # Westinghouse SOP tablosu -> SADECE west_map doluysa doldur
        if west_map:
            self._fill_west_sop_from_westhouse(
                ws=ws,
                start_row=2,
                adim_sayisi=adim_sayisi,
                job_df=job_df,
                west_map=west_map
            )
        # west_map boşsa: Westinghouse SOP satırları boş kalacak (süre yazmayacağız)
            # MOST SOP tablosu -> most_map doluysa doldur
        if most_map:
            self._fill_most_sop_from_most(
                ws=ws,
                start_row=2,
                adim_sayisi=adim_sayisi,
                job_df=job_df,
                most_map=most_map
            )

        # SOP tabloları doldurulduktan sonra, her bandın yanına
        # VA–NVAN–NVA yüzde dağılım grafiklerini ekle
        self._add_sop_percent_charts(ws, start_row=2, adim_sayisi=adim_sayisi)





        # CPM + Gantt blokları SOP'tan sonra başlasın
        top_row = sop_bottom + 2
        left_col = 1     # Zaman Etüdü
        mid_col = 11     # Westinghouse
        right_col = 21   # MOST


        cpm_time = self._build_cpm_rows_linear(steps_df, zaman_map)
        cpm_most = self._build_cpm_rows_linear(steps_df, most_map)

        if west_map:
            cpm_west = self._build_cpm_rows_linear(steps_df, west_map)
        else:
            cpm_west = []  # Westinghouse verisi yoksa CPM+Gantt çizme

        end1 = self._write_cpm_and_gantt_block(
            ws, start_row=top_row, col_start=left_col,
            title="Zaman Etüdü - CPM + Gantt", cpm_rows=cpm_time
        )
        end2 = self._write_cpm_and_gantt_block(
            ws, start_row=top_row, col_start=mid_col,
            title="Westinghouse - CPM + Gantt", cpm_rows=cpm_west
        )
        end3 = self._write_cpm_and_gantt_block(
            ws, start_row=top_row, col_start=right_col,
            title="MOST - CPM + Gantt", cpm_rows=cpm_most
        )

        base_row = max(end1, end2, end3) + 25

        # -------- Altta MEVCUT ÇIKTILAR (başlıklar aynı) --------

        # SOL: Westinghouse blokları (A..G) – sadece başlangıç satırını aşağı alıyoruz
        wh_start_row = base_row
        west_for_job = pd.DataFrame()
        if not self.west.empty:
            # Önce AdımID ile eşle (daha güvenli)
            if "AdımID" in self.west.columns:
                west_for_job = self.west[self.west["AdımID"].isin(adim_ids)].copy()
            # Hâlâ boşsa MOST AnalizID kesişimini dene
            if west_for_job.empty and "AnalizID" in self.west.columns:
                analiz_ids = ilgili_analizler["AnalizID"].dropna().unique().tolist()
                west_for_job = self.west[self.west["AnalizID"].isin(analiz_ids)].copy()

        if not west_for_job.empty:
            for aid, g in west_for_job.groupby("AnalizID", sort=False):
                # Adım adını bul (önce MOST'tan, yoksa iş adımlarından)
                step_nm = ""
                if (merged["AnalizID"] == aid).any():
                    s = merged.loc[merged["AnalizID"] == aid, "Adım Adı"].dropna()
                    if not s.empty:
                        step_nm = s.iloc[0]
                if not step_nm:
                    if "AdımID" in g.columns:
                        gid = g.iloc[0].get("AdımID")
                        if pd.notna(gid):
                            s2 = job_df.loc[job_df["AdımID"] == gid, "Adım Adı"]
                            if not s2.empty:
                                step_nm = s2.iloc[0]
                if not step_nm and not job_df.empty:
                    step_nm = job_df["Adım Adı"].dropna().iloc[0]

                wh_start_row = self._write_west_block(ws, wh_start_row, step_nm, g.iloc[0])

        # SAĞ: MOST tablosu (J..P) – başlıklar aynı, sadece daha aşağıdan başlıyor
        if not ilgili_analizler.empty:
            self._write_most_table(ws, start_row=base_row, merged=merged)

        # Otomatik genişlik
        self._autosize(ws)


    # ------------ Dışa Açık: Raporu Oluştur ------------

    from pathlib import Path
    from openpyxl import Workbook

    def build(self, output_path: str) -> str:
        """
        output_path: üretilecek xlsx dosya yolu
        """
        # 1) CSV'leri yükle
        self._load()

        # 2) Eğer sadece belirli JobID'ler seçildiyse, İş tablosunu filtrele
        if self.selected_job_ids and "JobID" in self.is_adimlari.columns:
            self.is_adimlari = self.is_adimlari[
                self.is_adimlari["JobID"].isin(self.selected_job_ids)
            ]

        wb = Workbook()
        # Varsayılan boş sayfayı kaldır
        wb.remove(wb.active)

        if self.is_adimlari.empty:
            ws = wb.create_sheet("Rapor")
            ws.cell(row=1, column=1, value="Veri bulunamadı.")
        else:
            # Eğer Ana Sayfa fonksiyonun varsa önce onu yaz
            if hasattr(self, "_write_main_sheet"):
                self._write_main_sheet(wb)

            # Sonra sadece filtrelenmiş İşler için sayfa oluştur
            for job_name, job_df in self.is_adimlari.groupby("İş Adı"):
                self._write_sheet_for_job(wb, job_name, job_df)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path



# ------------------------- Kullanım / Entegrasyon -------------------------

def create_most_report_job_based(
    output_path: str,
    paths: Dict[str, str],
    selected_job_ids=None
) -> str:
    """
    Uygulama içinden tek satırla çağır:
        create_most_report_job_based("cikti/Most_Rapor_IsBazli.xlsx", {
            "is_adimlari": ".../Is_Adimlari.csv",
            "zaman_etudu": ".../Zaman_Etudu.csv",
            "basic_analiz": ".../Basic_Most_Analizleri.csv",
            "basic_detay":  ".../Basic_Most_Detaylari.csv",
            "maxi_analiz":  ".../Maxi_Most_Analizleri.csv",
            "maxi_detay":   ".../Maxi_Most_Detaylari.csv",
            "mini_analiz":  ".../Mini_Most_Analizleri.csv",
            "mini_detay":   ".../Mini_Most_Detaylari.csv",
            "westinghouse": ".../Westinghouse_Analizleri.csv",
        }, selected_job_ids=[1, 3])
    """
    reporter = MostExcelJobReport(paths, selected_job_ids=selected_job_ids)
    return reporter.build(output_path)

